import csv
import matplotlib.pyplot as plt
import numpy as np
import pathlib
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side


class Vacancy:
    inRubles = {"KGS": 0.76, "BYR": 23.91, "GEL": 21.74, "EUR": 59.90, "AZN": 35.68,
                "KZT": 0.13, "UAH": 1.64, "RUR": 1, "USD": 60.66, "UZS": 0.0055,
                }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salaryTo = int(float(vacancy['salary_to']))
        self.salaryFrom = int(float(vacancy['salary_from']))
        self.salaryCurrency = vacancy['salary_currency']
        self.salaryAverage = ((self.salaryFrom + self.salaryTo) * self.inRubles[self.salaryCurrency]) / 2
        self.areaName = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.fileName = file_name
        self.vacancyName = vacancy_name

    @staticmethod
    def increment(dict, key, quantity):
        if key in dict:
            dict[key] += quantity
        else:
            dict[key] = quantity

    @staticmethod
    def average(d):
        dictionary = {}
        for key, values in d.items():
            dictionary[key] = int(sum(values)) / int(len(values))
        return dictionary

    def csv_reader(self):
        with open(self.fileName, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            headerLength = len(header)
            for i in reader:
                if len(i) == headerLength and '' not in i:
                    yield dict(zip(header, i))

    def get_statistic(self):
        salary = {}
        salaryVacancy = {}
        salaryCity = {}
        count = 0
        for vacancyDict in self.csv_reader():
            vacancy = Vacancy(vacancyDict)
            self.increment(salary, vacancy.year, [vacancy.salaryAverage])
            if vacancy.name.find(self.vacancyName) != -1:
                self.increment(salaryVacancy, vacancy.year, [vacancy.salaryAverage])
            self.increment(salaryCity, vacancy.areaName, [vacancy.salaryAverage])
            count += 1
        number = dict([(key, len(value)) for key, value in salary.items()])
        numbersName = dict([(key, len(value)) for key, value in salaryVacancy.items()])
        if not salaryVacancy:
            salaryVacancy = dict([(key, [0]) for key, value in salary.items()])
            numbersName = dict([(key, 0) for key, value in number.items()])
        statistics = self.average(salary)
        statistics2 = self.average(salaryVacancy)
        statistics3 = self.average(salaryCity)
        statistics4 = {}
        for year, salaries in salaryCity.items():
            statistics4[year] = round(len(salaries) / count, 4)
        statistics4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in statistics4.items()]))
        statistics4.sort(key=lambda a: a[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda a: a[0] in list(statistics4.keys()), [(key, value) for key, value
                                                                               in statistics3.items()]))
        statistics3.sort(key=lambda a: a[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])
        return statistics, number, statistics2, numbersName, statistics3, statistics5

    @staticmethod
    def print_statistic(statistics, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))


class Report:
    def __init__(self, vacancyName, statistics, statistics2, statistics3, statistics4, statistics5, statistics6):
        self.workbook = Workbook()
        self.vacancyName = vacancyName
        self.statistics = statistics
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    def get_excel(self):
        c = self.workbook.active
        c.title = 'Статистика по годам'
        c.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
                  'Количество вакансий - ' + self.vacancyName])
        for year in self.statistics.keys():
            c.append([year, self.statistics[year], self.statistics3[year], self.statistics2[year],
                      self.statistics4[year]])
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        columnWidth = []
        for x in data:
            for i, cell in enumerate(x):
                if len(columnWidth) > i:
                    if len(cell) > columnWidth[i]:
                        columnWidth[i] = len(cell)
                else:
                    columnWidth += [len(cell)]
        for i, columnWidth in enumerate(columnWidth, 1):
            c.column_dimensions[get_column_letter(i)].width = 2 + columnWidth
        newData = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city, value), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()):
            newData.append([city, value, '', city2, value2])
        d = self.workbook.create_sheet('Статистика по городам')
        for x in newData:
            d.append(x)
        columnWidth = []
        for x in newData:
            for i, cellule in enumerate(x):
                cellule = str(cellule)
                if len(columnWidth) > i:
                    if len(cellule) > columnWidth[i]:
                        columnWidth[i] = len(cellule)
                else:
                    columnWidth += [len(cellule)]
        for i, columnWidth in enumerate(columnWidth, 1):
            d.column_dimensions[get_column_letter(i)].width = columnWidth + 2
        fontBold = Font(bold=True)
        for column in 'ABCDE':
            c[column + '1'].font = fontBold
            d[column + '1'].font = fontBold
        for index, _ in enumerate(self.statistics5):
            d['E' + str(2 + index)].number_format = '0.00%'
        fontThin = Side(border_style='thin', color='00000000')
        for x in range(len(newData)):
            for column in 'ABDE':
                d[column + str(1 + x)].border = Border(left=fontThin, bottom=fontThin, right=fontThin, top=fontThin)
        for x, _ in enumerate(self.statistics):
            for column in 'ABCDE':
                c[column + str(1 + x)].border = Border(left=fontThin, bottom=fontThin, right=fontThin, top=fontThin)
        self.workbook.save(filename='report.xlsx')

    def get_image(self):
        fig, ((axes, axes2), (axes3, axes4)) = plt.subplots(nrows=2, ncols=2)
        bar = axes.bar(np.array(list(self.statistics.keys())) - 0.4, self.statistics.values(), width=0.4)
        bar2 = axes.bar(np.array(list(self.statistics.keys())), self.statistics3.values(), width=0.4)
        axes.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        axes.grid(axis='y')
        axes.legend((bar[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancyName.lower()), prop={'size': 8})
        axes.set_xticks(np.array(list(self.statistics.keys())) - 0.2, list(self.statistics.keys()), rotation=90)
        axes.xaxis.set_tick_params(labelsize=8)
        axes.yaxis.set_tick_params(labelsize=8)
        axes2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar = axes2.bar(np.array(list(self.statistics2.keys())) - 0.4, self.statistics2.values(), width=0.4)
        bar2 = axes2.bar(np.array(list(self.statistics2.keys())), self.statistics4.values(), width=0.4)
        axes2.legend((bar[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancyName.lower()),
                     prop={'size': 8})
        axes2.set_xticks(np.array(list(self.statistics2.keys())) - 0.2, list(self.statistics2.keys()), rotation=90)
        axes2.grid(axis='y')
        axes2.xaxis.set_tick_params(labelsize=8)
        axes2.yaxis.set_tick_params(labelsize=8)
        axes3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        axes3.barh(
            list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.statistics5.keys()))]),
            list(reversed(list(self.statistics5.values()))), color='blue', height=0.5, align='center')
        axes3.yaxis.set_tick_params(labelsize=6)
        axes3.xaxis.set_tick_params(labelsize=8)
        axes3.grid(axis='x')
        axes4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        e = 1 - sum([value for value in self.statistics6.values()])
        axes4.pie(list(self.statistics6.values()) + [e], labels=list(self.statistics6.keys()) + ['Другие'],
                  textprops={'fontsize': 6})
        plt.tight_layout()
        plt.savefig('graph.png')

    def get_pdf(self):
        statistic = []
        environment = Environment(loader=FileSystemLoader('.'))
        template = environment.get_template("html.html")
        for i in self.statistics.keys():
            statistic.append([i, self.statistics[i], self.statistics2[i], self.statistics3[i],
                              self.statistics4[i]])
        for x in self.statistics6:
            self.statistics6[x] = round(100 * self.statistics6[x], 2)
        pdfTemp = template.render(
            {'name': self.vacancyName, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
             'stats': statistic, 'stats5': self.statistics5, 'stats6': self.statistics6})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdfTemp, 'report.pdf', configuration=config, options={"enable-local-file-access": None})


class InputConnect:
    def __init__(self):
        self.fileName = input('Введите название файла: ')
        self.vacancyName = input('Введите название профессии: ')
        a = DataSet(self.fileName, self.vacancyName)
        statistics, statistics2, statistics3, statistics4, statistics5, statistics6 = a.get_statistic()
        a.print_statistic(statistics, statistics2, statistics3, statistics4, statistics5, statistics6)
        b = Report(self.vacancyName, statistics, statistics2, statistics3, statistics4, statistics5, statistics6)
        b.get_excel()
        b.get_image()
        b.get_pdf()


def get_answer():
    InputConnect()
